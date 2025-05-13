from typing import Any, List, Type, Dict, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sheetdb.logger_utils import get_logger
from sheetdb.interfaces.base import BaseSheetDB
from sheetdb.interfaces.model import SheetModel
from sheetdb.exceptions import NotFoundError

logger = get_logger("exceldb")

cache_headers: Dict[str, List[str]] = {}


class ExcelDB(BaseSheetDB):
    def __init__(self, file_path: str):
        self.file_path = file_path
        self._workbook = None

    def _get_workbook(self):
        if not self._workbook:
            try:
                self._workbook = load_workbook(self.file_path)
            except FileNotFoundError:
                self._workbook = Workbook()
                self._workbook.save(self.file_path)
        return self._workbook

    def _get_worksheet(
        self, workbook, sheet_name: str, headers: List[str] = None
    ) -> Worksheet:
        if sheet_name not in workbook.sheetnames:
            if headers is None:
                raise ValueError(
                    f"Sheet '{sheet_name}' not found and no headers provided to create it."
                )
            ws = workbook.create_sheet(sheet_name)
            ws.append(headers)
            cache_headers[sheet_name] = headers
            workbook.save(self.file_path)
            logger.info(f"Created new worksheet: {sheet_name} with headers: {headers}")
            return ws
        else:
            return workbook[sheet_name]

    def _get_headers(self, ws, sheet_name):
        if sheet_name in cache_headers:
            return cache_headers[sheet_name]
        headers = [cell.value for cell in ws[1]]
        cache_headers[sheet_name] = headers
        return headers

    def insert(self, model: SheetModel):
        self.insert_many([model])

    def insert_many(self, models: List[SheetModel]):
        if not models:
            return

        workbook = self._get_workbook()
        sheet_name = models[0].get_sheet_name()

        # Получаем field_map для экземпляра модели
        field_map = models[0].get_field_map()
        if not isinstance(field_map, dict):
            raise ValueError(f"Invalid field_map for {models[0]}: {field_map}")

        mapped_rows = []
        for model in models:
            raw = {field_map.get(k, k): v for k, v in model.model_dump().items()}
            mapped_rows.append(raw)

        ws = self._get_worksheet(
            workbook, sheet_name, headers=list(mapped_rows[0].keys())
        )
        headers = self._get_headers(ws, sheet_name)

        for row in mapped_rows:
            row_values = [row.get(header, "") for header in headers]
            ws.append(row_values)

        workbook.save(self.file_path)
        logger.info(f"Inserted {len(mapped_rows)} row(s) into {sheet_name}")

    def get_all(
        self, model: Type[SheetModel], filters: Optional[Dict[str, Any]] = None
    ):
        workbook = self._get_workbook()
        sheet_name = model.get_class_sheet_name()
        ws = self._get_worksheet(workbook, sheet_name)

        headers = self._get_headers(ws, sheet_name)
        rows = list(ws.iter_rows(min_row=2, values_only=True))  # Пропускаем заголовок

        # Получаем маппинг для переворота
        field_map = model.generate_field_map()

        data = []
        for row in rows:
            # Преобразуем полученные данные с учетом маппинга полей
            model_data = {
                field_map.get(header, header): value
                for header, value in zip(headers, row)
            }

            # Создаем объект модели с преобразованными данными
            try:
                item = model(**model_data)
            except Exception as e:
                logger.error(f"Failed to parse row {row}: {e}")
                continue

            # Фильтрация по условиям
            if filters:
                match = True
                for key, value in filters.items():
                    if getattr(item, key, None) != value:
                        match = False
                        break
                if match:
                    data.append(item)
            else:
                data.append(item)

        return data

    def get_one(self, model: Type[SheetModel], filters: Dict[str, str]):
        """Получить одну запись по фильтру."""
        data = self.get_all(model)

        for item in data:
            match = True
            for key, value in filters.items():
                if getattr(item, key, None) != value:
                    match = False
                    break
            if match:
                return item

        raise NotFoundError(f"Record not found for filters: {filters}")

    def update(
        self,
        model: Type[SheetModel],
        filters: Dict[str, str],
        update_data: Dict[str, str],
    ):
        """Обновить запись в Excel."""
        data = self.get_all(model)

        for idx, item in enumerate(data):
            if all(getattr(item, k, None) == v for k, v in filters.items()):
                # Применяем обновления к объекту
                for k, v in update_data.items():
                    setattr(item, k, v)

                workbook = self._get_workbook()
                sheet_name = model.get_class_sheet_name()
                ws = self._get_worksheet(workbook, sheet_name)
                headers = self._get_headers(ws, sheet_name)

                # Подготовка данных для обновления
                mapped_row = []
                for header in headers:
                    field_name = model.generate_field_map().get(header, header)
                    value = getattr(item, field_name, "")
                    mapped_row.append(value)

                row_number = (
                    idx + 2
                )  # Строка для обновления (плюс 2: одна строка заголовков + индекс с 0)
                for col_num, value in enumerate(mapped_row, 1):
                    ws.cell(row=row_number, column=col_num, value=value)

                workbook.save(self.file_path)
                logger.info(f"Updated record in {sheet_name} at row {row_number}")
                return item

        raise NotFoundError(f"Record not found for filters: {filters}")

    def delete(self, model: Type[SheetModel], filters: Dict[str, str]):
        """Удалить запись из базы данных по фильтру."""
        data = self.get_all(model)

        for item in data:
            match = True
            for key, value in filters.items():
                if getattr(item, key, None) != value:
                    match = False
                    break
            if match:
                self.delete_row(model, item)
                return item

        raise NotFoundError(f"Record not found for filters: {filters}")

    def delete_row(self, model: Type[SheetModel], item: SheetModel):
        """Удалить строку из Excel."""
        workbook = self._get_workbook()
        sheet_name = model.get_class_sheet_name()
        ws = self._get_worksheet(workbook, sheet_name)

        # Находим строку с данными
        row_number = int(item.id) + 1  # Сдвигаем на 1, чтобы пропустить заголовок
        ws.delete_rows(row_number)  # Удаляем строку с данными

        workbook.save(self.file_path)
        logger.info(f"Deleted row {row_number} in {sheet_name}")

    def delete_all(self, model: Type[SheetModel]):
        """Быстро очистить все записи на листе, кроме заголовка."""
        workbook = self._get_workbook()
        sheet_name = model.get_class_sheet_name()
        ws = self._get_worksheet(workbook, sheet_name)

        rows = list(ws.iter_rows(min_row=2))  # Пропускаем заголовок
        if not rows:
            logger.info(f"No data to clear on sheet '{sheet_name}'")
            return

        # Определяем диапазон строк, которые нужно очистить
        ws.delete_rows(2, len(rows))  # Удаляем все строки

        workbook.save(self.file_path)
        logger.info(f"Cleared {len(rows)} row(s) from sheet '{sheet_name}'")
